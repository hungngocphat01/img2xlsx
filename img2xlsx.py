import os
import argparse
import openpyxl
import pandas as pd

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

from PIL import Image
from tqdm import tqdm


def img_pixel_hex(img: Image, x: int, y: int):
    pixel_index = img.getpixel((x, y))
    rgb_value = img.getpalette()[pixel_index * 3:pixel_index * 3 + 3]
    return "{:02x}{:02x}{:02x}".format(rgb_value[0], rgb_value[1], rgb_value[2]).upper()


def set_sheet_col_width(ws: Worksheet, width: int):
    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=width)
    ws.column_dimensions = dim_holder


def fill_sheet(ws: Worksheet, img: Image):
    ncols, nrows = img.size

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=nrows, min_col=1, max_col=ncols)):
        for j, cell in enumerate(row):
            hex_color = img_pixel_hex(img, j, i)
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def start_draw(frames_dir: str, pal_img: Image, dither=0, colwith=2.25, zoomscale=25, head: int = None):
    """
    Args:
        frames_dir: directory containing the frames. Filename should have the format of %05d.jpg
        pal_img: dummy palette image
        dither: whether to dither. Defaults to no dither. See PIL.Image.quantize for more details
        colwidth: with of the columns. One unit equals 7 pixels (as of writing this)
        zoomscale: a number from 0-100 specifying the default zoom level of the sheet
        head: generate only this amount of images instead of the whole input directory
    """

    wb = openpyxl.Workbook()

    MAX_LEN = len(os.listdir(frames_dir))
    for i in tqdm(range(1, head or MAX_LEN)):
        ws = wb.create_sheet(f"{i:04d}")

        img = Image.open(os.path.join(frames_dir, f"{i:04d}.jpg"))
        img = img.quantize(palette=pal_img, dither=dither)

        fill_sheet(ws, img)
        set_sheet_col_width(ws, colwith)
        # why not setting row height as well? bcz i'm lazy

        if zoomscale is not None:
            ws.sheet_view.zoomScale = zoomscale

    return wb


def main(args):
    # X11 pallete, from https://www.ditig.com/256-colors-cheat-sheet
    color_df = pd.read_parquet("x11_palette.parquet")
    palette = [i for arr in color_df.RGB.to_list() for i in arr]

    # Dummy palette image
    pal_img = Image.new("P", (1, 1))
    pal_img.putpalette(palette)

    wb = start_draw(args.frames_dir, pal_img, 0, args.colwith, args.zoomscale, args.head)
    wb.save(args.output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--frames-dir",
        type=str,
        required=True,
        help=r"Path to directory containing frames. Image filenames should have the format %04d.jpg"
    )
    parser.add_argument("--output", type=str, required=True, help="Output file path")
    parser.add_argument("--head", type=int, help="Only convert the first N frames instead of the whole directory")
    parser.add_argument("--zoomscale", type=int, default=25, help="Default zoom scale of the output workbook")
    parser.add_argument("--colwidth", type=float, default=2.25, help="Column width of the output workbook")
    main(parser.parse_args())
